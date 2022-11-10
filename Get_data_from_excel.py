import openpyxl

b = openpyxl.load_workbook("D:/Python Project/WebAttack_using_POm/Setup_files/Data.xlsx")
sheet = b.active
unregistered_username = sheet.cell(row=1, column=2)
unregistered_password = sheet.cell(row=2, column=2)

new_user_email = sheet.cell(row=4, column=2)
new_user_password = sheet.cell(row=5, column=2)
new_user_first_name = sheet.cell(row=6, column=2)
new_user_last_name = sheet.cell(row=7, column=2)
new_user_current_address = sheet.cell(row=8, column=2)
new_user_address1 = sheet.cell(row=9, column=2)
new_user_zipcode = sheet.cell(row=10, column=2)
new_user_city = sheet.cell(row=11, column=2)
new_user_phone_number = sheet.cell(row=12, column=2)
new_user_state = sheet.cell(row=13, column=2)
